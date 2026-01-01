from PyQt5.QtWidgets import QWidget, QVBoxLayout, QHBoxLayout, QPushButton, QFileDialog, QTableWidgetItem, QApplication
from PyQt5.QtCore import Qt
from qfluentwidgets import TitleLabel, BodyLabel, TableWidget, PushButton, SubtitleLabel, SpinBox, ComboBox, InfoBar,InfoBarPosition, FluentIcon
import pyqtgraph as pg

# 延迟导入：这些库只在需要时才导入，加快应用启动速度
# import numpy as np
# from openpyxl import load_workbook, Workbook


class FunctionFitInterface(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        # 延迟导入标志
        self._libs_loaded = False
        self.setObjectName("functionFitInterface")


        main_layout = QHBoxLayout(self)
        main_layout.setSpacing(24)

        # 左侧：数据输入区
        left_layout = QVBoxLayout()
        left_layout.setSpacing(16)

        self.dataTable = TableWidget(self)
        self.dataTable.setColumnCount(2)
        self.dataTable.setHorizontalHeaderLabels(["x", "y"])
        self.dataTable.setColumnWidth(0, 125)
        self.dataTable.setColumnWidth(1, 125)
        left_layout.addWidget(self.dataTable)

        btn_layout = QHBoxLayout()
        add_row_btn = PushButton("添加一行")
        add_row_btn.clicked.connect(self.add_row)

        del_row_btn = PushButton("删除选中行")  # 新增按钮
        del_row_btn.clicked.connect(self.delete_selected_row)  # 绑定槽函数
        btn_layout.addWidget(add_row_btn)

        btn_layout.addWidget(del_row_btn)  # 添加到布局
        left_layout.addLayout(btn_layout)



        btn_layout = QHBoxLayout()
        import_btn = PushButton("导入 Excel")
        import_btn.clicked.connect(self.import_excel)
        export_btn = PushButton("导出 Excel")
        export_btn.clicked.connect(self.export_excel)
        btn_layout.addWidget(import_btn)
        btn_layout.addWidget(export_btn)
        left_layout.addLayout(btn_layout)

        self.dataTable.setMinimumWidth(280)
        self.dataTable.setMaximumWidth(280)
        main_layout.addLayout(left_layout, 1)

        self.add_row()

        # 右侧：图像展示区
        right_layout = QVBoxLayout()
        right_layout.setSpacing(12)
        right_layout.addWidget(SubtitleLabel("函数图像预览"))

        # 占位符，实际的canvas会在_load_heavy_libraries中创建
        self.canvas_placeholder = QWidget()
        self.canvas_layout = QVBoxLayout(self.canvas_placeholder)
        self.canvas_layout.setContentsMargins(0, 0, 0, 0)
        right_layout.addWidget(self.canvas_placeholder, stretch=1)

        self.resultLabel = BodyLabel("")
        self.resultLabel.setWordWrap(True)  # 自动换行
        right_layout.addWidget(self.resultLabel)


        # 拟合阶数和输出语言选择（合并到同一行）
        options_layout = QHBoxLayout()
        self.spinBox = SpinBox()
        self.spinBox.setRange(1, 10)
        self.spinBox.setValue(2)
        options_layout.addWidget(SubtitleLabel("拟合阶数"))
        options_layout.addWidget(self.spinBox)

        self.langBox = ComboBox()
        self.langBox.addItems(["C/C++", "Python"])
        options_layout.addWidget(SubtitleLabel("输出语言"))
        options_layout.addWidget(self.langBox)


        right_layout.addLayout(options_layout)


        # 代码显示和复制按钮
        self.codeLabel = BodyLabel("")
        self.codeLabel.setWordWrap(True)  # 自动换行
        right_layout.addWidget(self.codeLabel)

        btn_layout = QHBoxLayout()  # 新增一行布局
        fit_btn = PushButton(FluentIcon.UNIT,"拟合并绘图")
        fit_btn.clicked.connect(self.fit_and_plot)
        btn_layout.addWidget(fit_btn)
        copy_btn = PushButton(FluentIcon.COPY, "复制代码")
        copy_btn.clicked.connect(self.copy_code)
        btn_layout.addWidget(copy_btn)
        right_layout.addLayout(btn_layout)


        main_layout.addLayout(right_layout, 2)

    def _load_heavy_libraries(self):
        """延迟加载大型库，提高应用启动速度"""
        if self._libs_loaded:
            return
        
        global np, load_workbook, Workbook
        
        import numpy as np
        from openpyxl import load_workbook, Workbook
        
        # 创建 PyQtGraph 画布
        self.plot_widget = pg.PlotWidget()
        self.plot_widget.setBackground('w')  # 白色背景
        self.plot_widget.showGrid(x=True, y=True, alpha=0.3)
        self.plot_widget.setLabel('left', 'y')
        self.plot_widget.setLabel('bottom', 'x')
        self.plot_widget.setTitle('graph of a function')
        
        # 将 plot_widget 添加到占位符布局中
        if hasattr(self, 'canvas_layout'):
            self.canvas_layout.addWidget(self.plot_widget)
        
        self._libs_loaded = True

    def add_row(self):
        row = self.dataTable.rowCount()
        self.dataTable.insertRow(row)
        # 可选：初始化为空字符串
        self.dataTable.setItem(row, 0, QTableWidgetItem(""))
        self.dataTable.setItem(row, 1, QTableWidgetItem(""))

    def delete_selected_row(self):
        selected = self.dataTable.selectedItems()
        if selected:
            rows = set(item.row() for item in selected)
            for row in sorted(rows, reverse=True):
                self.dataTable.removeRow(row)

    def import_excel(self):
        self._load_heavy_libraries()  # 延迟加载库
        path, _ = QFileDialog.getOpenFileName(self, "导入 Excel", "", "Excel Files (*.xlsx)")
        if path:
            wb = load_workbook(path)
            ws = wb.active
            self.dataTable.setRowCount(0)
            for row_data in ws.iter_rows(min_row=2, values_only=True):  # 跳过表头
                row = self.dataTable.rowCount()
                self.dataTable.insertRow(row)
                for col, value in enumerate(row_data[:2]):
                    item = QTableWidgetItem(str(value) if value is not None else "")
                    self.dataTable.setItem(row, col, item)


    def export_excel(self):
        self._load_heavy_libraries()  # 延迟加载库
        path, _ = QFileDialog.getSaveFileName(self, "导出 Excel", "", "Excel Files (*.xlsx)")
        if path:
            data = self.parse_data()
            if data is not None:
                wb = Workbook()
                ws = wb.active
                ws.append(["x", "y"])
                for row in data:
                    ws.append(row)
                wb.save(path)

    def parse_data(self):
        data = []
        row_count = self.dataTable.rowCount()
        for row in range(row_count):
            try:
                x_item = self.dataTable.item(row, 0)
                y_item = self.dataTable.item(row, 1)
                if x_item is None or y_item is None:
                    continue
                x = float(x_item.text())
                y = float(y_item.text())
                data.append([x, y])
            except Exception:
                continue
        return data if data else None

    def fit_and_plot(self):
        self._load_heavy_libraries()  # 延迟加载库
        data = self.parse_data()
        if not data:
            self.resultLabel.setText("数据格式错误或为空")
            self.codeLabel.setText("")
            return

        x = np.array([d[0] for d in data])
        y = np.array([d[1] for d in data])
        degree = self.spinBox.value()
        coeffs = np.polyfit(x, y, degree)

        # 用更密集的横坐标画拟合曲线
        x_fit = np.linspace(x.min(), x.max(), 100)
        y_fit = np.polyval(coeffs, x_fit)

        # 清空并重新绘图
        self.plot_widget.clear()
        
        # 绘制原始数据点（蓝色散点）
        scatter = pg.ScatterPlotItem(
            x=x, y=y, 
            pen=None, 
            brush=pg.mkBrush(0, 0, 255, 120),  # 蓝色半透明
            size=10,
            name='raw data'
        )
        self.plot_widget.addItem(scatter)
        
        # 绘制拟合曲线（红色线条）
        pen = pg.mkPen(color=(255, 0, 0), width=2)  # 红色线条
        curve = self.plot_widget.plot(
            x_fit, y_fit, 
            pen=pen,
            name='Fitted curve'
        )

        # 添加图例
        self.plot_widget.addLegend()

        formula = self.poly_formula(coeffs)
        self.resultLabel.setText(f"拟合公式: {formula}")

        lang = self.langBox.currentText()
        code = self.generate_code(coeffs, lang)
        self.codeLabel.setText(code)

    def poly_formula(self, coeffs):
        terms = []
        degree = len(coeffs) - 1
        for i, c in enumerate(coeffs):
            power = degree - i
            if abs(c) < 1e-8:
                continue
            if power == 0:
                terms.append(f"{c:.6g}")
            elif power == 1:
                terms.append(f"{c:.6g}*x")
            else:
                terms.append(f"{c:.6g}*x^{power}")
        return " + ".join(terms)

    def generate_code(self, coeffs, lang):
        degree = len(coeffs) - 1
        if lang == "C/C++":
            code = "double poly(double x) {\n    return "
        elif lang == "Python":
            code = "def poly(x):\n    return "
        else:
            code = ""
        terms = []
        for i, c in enumerate(coeffs):
            power = degree - i
            if abs(c) < 1e-8:
                continue
            if power == 0:
                terms.append(f"{c:.6g}")
            elif power == 1:
                terms.append(f"{c:.6g}*x")
            else:
                terms.append(f"{c:.6g}*pow(x,{power})" if lang == "C/C++" else f"{c:.6g}*x**{power}")
        code += " + ".join(terms)
        code += ";\n}" if lang == "C/C++" else ""
        return code


    def copy_code(self):
        clipboard = QApplication.clipboard()
        clipboard.setText(self.codeLabel.text())
        # 弹出提示
        InfoBar.success(
            title='复制成功',
            content="代码已复制到剪贴板！",
            orient=Qt.Horizontal,
            isClosable=True,
            position=InfoBarPosition.TOP,
            duration=2000,
            parent=self
        )